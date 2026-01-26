"""Excel/CSV processor using pandas and openpyxl"""

import io
import re
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl

from .base import BaseProcessor


class ExcelProcessor(BaseProcessor):
    """Processor for Excel and CSV files using direct parsing"""
    
    SUPPORTED_TYPES = ['xlsx', 'xls', 'csv']
    
    def supports(self, file_type: str) -> bool:
        """Check if this processor supports the file type."""
        return file_type.lower() in self.SUPPORTED_TYPES
    
    def process(self, file_content: bytes, file_type: str,
               metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Process Excel/CSV file and extract text content.
        
        Args:
            file_content: File content as bytes
            file_type: File extension
            metadata: Optional file metadata
        
        Returns:
            Dictionary with 'raw_text' and 'metadata' keys
        """
        file_type_lower = file_type.lower()
        
        try:
            if file_type_lower == 'csv':
                # Read CSV
                df = pd.read_csv(io.BytesIO(file_content))
            elif file_type_lower == 'xlsx':
                # Read XLSX
                df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
            elif file_type_lower == 'xls':
                # Read XLS (older format)
                df = pd.read_excel(io.BytesIO(file_content), engine='xlrd')
            else:
                raise ValueError(f"Unsupported Excel file type: {file_type}")
            
            # Convert DataFrame to text representation
            # Include column names and all data
            text_parts = []
            
            # Add column names
            text_parts.append("Columns: " + ", ".join(df.columns.astype(str)))
            text_parts.append("\n")
            
            # Add data rows
            for idx, row in df.iterrows():
                row_text = " | ".join([f"{col}: {val}" for col, val in row.items()])
                text_parts.append(f"Row {idx + 1}: {row_text}")
            
            raw_text = "\n".join(text_parts)
            
            # Create metadata
            processing_metadata = {
                'file_type': file_type,
                'processing_method': 'pandas_direct_parse',
                'num_rows': len(df),
                'num_columns': len(df.columns),
                'column_names': df.columns.tolist(),
                'dataframe_shape': list(df.shape)
            }
            
            return {
                'raw_text': raw_text,
                'metadata': processing_metadata
            }
        
        except Exception as e:
            raise Exception(f"Error processing Excel/CSV file: {str(e)}")
    
    def extract_data_between_delimiters(
        self, 
        file_content: bytes, 
        file_type: str,
        delimiter_pattern: str = r'\*{4,}'  # 4+ asterisks
    ) -> pd.DataFrame:
        """Extract DataFrame containing only rows between delimiter rows.
        
        Args:
            file_content: File content as bytes
            file_type: File extension
            delimiter_pattern: Regex pattern for delimiter rows (default: 4+ asterisks)
        
        Returns:
            DataFrame with data rows only (header row + data rows between delimiters)
        
        Raises:
            ValueError: If delimiter rows are not found or data section is empty
        """
        file_type_lower = file_type.lower()
        
        try:
            # Read Excel file without header initially
            if file_type_lower == 'csv':
                df = pd.read_csv(io.BytesIO(file_content), header=None)
            elif file_type_lower == 'xlsx':
                df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl', header=None)
            elif file_type_lower == 'xls':
                df = pd.read_excel(io.BytesIO(file_content), engine='xlrd', header=None)
            else:
                raise ValueError(f"Unsupported Excel file type: {file_type}")
            
            # Find delimiter rows
            delimiter_rows = []
            for idx, row in df.iterrows():
                if self._is_delimiter_row(row, delimiter_pattern):
                    delimiter_rows.append(idx)
            
            if len(delimiter_rows) < 3:
                raise ValueError(
                    f"Could not find enough delimiter rows (expected at least 3, found {len(delimiter_rows)}). "
                    f"Delimiter pattern: {delimiter_pattern}. Found at rows: {delimiter_rows}"
                )
            
            # Structure: Delimiter 1 -> Header -> Delimiter 2 -> Data -> Delimiter 3 -> Summary
            # Delimiter 1 marks start
            delimiter1_idx = delimiter_rows[0]
            # Delimiter 2 is after header (separator)
            delimiter2_idx = delimiter_rows[1]
            # Delimiter 3 marks end of data (second-to-last delimiter)
            delimiter3_idx = delimiter_rows[-2]  # Second-to-last delimiter
            
            # Header row is between delimiter 1 and 2
            header_row_idx = delimiter1_idx + 1
            
            if header_row_idx >= delimiter2_idx:
                raise ValueError(
                    f"Header row index ({header_row_idx}) is at or after delimiter 2 ({delimiter2_idx})"
                )
            
            # Data starts after delimiter 2
            data_start_idx = delimiter2_idx + 1
            
            # Data ends at delimiter 3
            data_end_idx = delimiter3_idx
            
            if data_start_idx >= data_end_idx:
                raise ValueError(
                    f"Data start index ({data_start_idx}) is at or after end delimiter ({data_end_idx})"
                )
            
            # Extract header row
            header_row = df.iloc[header_row_idx]
            
            # Extract data section (rows between data_start and data_end)
            data_section = df.iloc[data_start_idx:data_end_idx].copy()
            
            if len(data_section) == 0:
                raise ValueError("Data section between delimiters is empty")
            
            # Set header row as column names
            data_section.columns = header_row.values
            
            # Remove completely empty rows
            data_section = data_section.dropna(how='all')
            
            if len(data_section) == 0:
                raise ValueError("No data rows found after removing empty rows")
            
            return data_section
        
        except Exception as e:
            raise Exception(f"Error extracting data between delimiters: {str(e)}")
    
    def _is_delimiter_row(self, row: pd.Series, pattern: str) -> bool:
        """Check if row is a delimiter row (all cells match pattern).
        
        Args:
            row: pandas Series representing a row
            pattern: Regex pattern to match (e.g., r'\*{4,}' for 4+ asterisks)
        
        Returns:
            True if row is a delimiter row, False otherwise
        """
        # Get all non-null cell values as strings
        cell_values = [str(val) for val in row.values if pd.notna(val)]
        
        if len(cell_values) == 0:
            return False
        
        # Check if all cells match the pattern (contain asterisks)
        pattern_re = re.compile(pattern)
        for cell_value in cell_values:
            if not pattern_re.search(cell_value):
                return False
        
        return True
    
    def normalize_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert Excel column names to snake_case for DB compatibility.
        
        Converts:
        - "Withdrawal Amt." → "withdrawal_amt"
        - "Chq./Ref.No." → "chq_ref_no"
        - "Value Dt" → "value_dt"
        - "Closing Balance" → "closing_balance"
        
        Args:
            df: DataFrame with original column names
        
        Returns:
            DataFrame with normalized column names
        """
        df_normalized = df.copy()
        
        # Normalize each column name
        new_columns = []
        for col in df.columns:
            # Convert to string and lowercase
            col_str = str(col).strip().lower()
            
            # Replace special characters and spaces with underscores
            # Remove dots, slashes, hyphens, etc.
            col_str = re.sub(r'[./\-]', '_', col_str)
            
            # Replace multiple spaces with single underscore
            col_str = re.sub(r'\s+', '_', col_str)
            
            # Remove any remaining special characters (keep only alphanumeric and underscore)
            col_str = re.sub(r'[^a-z0-9_]', '', col_str)
            
            # Remove multiple consecutive underscores
            col_str = re.sub(r'_+', '_', col_str)
            
            # Remove leading/trailing underscores
            col_str = col_str.strip('_')
            
            new_columns.append(col_str)
        
        df_normalized.columns = new_columns
        return df_normalized
